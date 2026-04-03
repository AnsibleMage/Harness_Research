"""Sheet 1: 표지 생성 — A4 가로, Navy/Blue 디자인"""
from openpyxl import Workbook
from .common import *

def create_cover_sheet(wb, meta):
    """표지 시트 생성. meta는 frontmatter 딕셔너리."""
    ws = wb.active
    ws.title = "표지"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    col_widths = {'A': 4, 'B': 4, 'C': 12, 'D': 8, 'E': 18, 'F': 18, 'G': 8, 'H': 12, 'I': 4, 'J': 4}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    for r in range(1, 26):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 10
    ws.row_dimensions[5].height = 12
    ws.row_dimensions[8].height = 50
    ws.row_dimensions[9].height = 8
    ws.row_dimensions[10].height = 38
    ws.row_dimensions[11].height = 15
    ws.row_dimensions[12].height = 3
    ws.row_dimensions[13].height = 30
    ws.row_dimensions[23].height = 30

    # 상단 바
    for col in range(1, 11):
        ws.cell(row=2, column=col).fill = FILL_NAVY

    # 기관명
    ws.merge_cells('C4:H4')
    c = ws.cell(row=4, column=3, value="서울특별시 정원도시국")
    c.font = FONT_ORG
    c.alignment = ALIGN_CENTER

    # 문서 제목
    ws.merge_cells('C8:H8')
    c = ws.cell(row=8, column=3, value=meta.get('title', ''))
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    # 사업명
    ws.merge_cells('C10:H10')
    project = meta.get('project', '')
    c = ws.cell(row=10, column=3, value=project.replace('·', '·\n') if '·' in project else project)
    c.font = FONT_SUBTITLE
    c.alignment = ALIGN_CENTER

    # 악센트 라인
    for col in range(4, 8):
        ws.cell(row=12, column=col).fill = FILL_ACCENT

    # 정보 테이블
    labels = ["문서번호", "버전", "작성일", "작성자", "검토자", "승인자", "보안등급"]
    keys = ["doc_id", "version", "created", "author", "reviewer", "approver", "classification"]

    for i, (lbl, key) in enumerate(zip(labels, keys)):
        row = 14 + i
        ws.row_dimensions[row].height = 28
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        lbl_cell = ws.cell(row=row, column=4, value=lbl)
        lbl_cell.font = FONT_LABEL
        lbl_cell.fill = FILL_ACCENT
        lbl_cell.alignment = ALIGN_CENTER

        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        val_cell = ws.cell(row=row, column=6, value=meta.get(key, ''))
        val_cell.font = FONT_VALUE
        val_cell.fill = FILL_WHITE if i % 2 == 0 else FILL_LIGHT
        val_cell.alignment = ALIGN_CENTER

    # 날짜
    ws.merge_cells('C22:H22')
    created = meta.get('created', '')
    date_str = created[:7].replace('-', '. ') + '.' if created else ''
    c = ws.cell(row=22, column=3, value=date_str)
    c.font = FONT_YEAR
    c.alignment = ALIGN_CENTER

    # 하단 바
    for col in range(1, 11):
        ws.cell(row=24, column=col).fill = FILL_NAVY

    ws.print_area = 'A1:J25'
    return ws
