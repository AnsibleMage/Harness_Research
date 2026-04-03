"""공통 스타일, 컬러, 폰트, 유틸리티 — req-definition-xlsx 스킬"""
import os, shutil, tempfile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# === 컬러 팔레트 (Navy 테마) ===
NAVY = "1B2A4A"
ACCENT_BLUE = "2E5090"
LIGHT_GRAY = "F2F4F7"
MID_GRAY = "D5D8DC"
BORDER_GRAY = "B0B8C4"
NOTE_GRAY = "5A6270"
WHITE = "FFFFFF"

# === 폰트 ===
FONT_TITLE = Font(name="맑은 고딕", size=22, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name="맑은 고딕", size=16, color=ACCENT_BLUE)
FONT_SHEET_TITLE = Font(name="맑은 고딕", size=18, bold=True, color=NAVY)
FONT_HEADER = Font(name="맑은 고딕", size=9, bold=True, color=WHITE)
FONT_HEADER_10 = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
FONT_DATA = Font(name="맑은 고딕", size=9, color=NAVY)
FONT_DATA_WRAP = Font(name="맑은 고딕", size=8.5, color=NAVY)
FONT_NO = Font(name="맑은 고딕", size=9, bold=True, color=ACCENT_BLUE)
FONT_NOTE = Font(name="맑은 고딕", size=8.5, color=NOTE_GRAY)
FONT_ORG = Font(name="맑은 고딕", size=10, color="7F8C8D")
FONT_YEAR = Font(name="맑은 고딕", size=13, color=ACCENT_BLUE, bold=True)
FONT_LABEL = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
FONT_VALUE = Font(name="맑은 고딕", size=11, color=NAVY)

# === 채우기 ===
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT_BLUE)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT_GRAY)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

# === 정렬 ===
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
ALIGN_LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)

# === 테두리 ===
THIN_SIDE = Side(style='thin', color=BORDER_GRAY)
BORDER_THIN = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BORDER_HEADER = Border(
    left=THIN_SIDE, right=THIN_SIDE,
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='medium', color=NAVY)
)

def safe_save(wb, target_path):
    """파일 잠금 대응: temp에 저장 후 복사. 실패 시 _v2 대체."""
    tmp = os.path.join(tempfile.gettempdir(), "req_def_temp.xlsx")
    wb.save(tmp)
    try:
        shutil.copy2(tmp, target_path)
        return target_path
    except PermissionError:
        alt = target_path.replace(".xlsx", "_v2.xlsx")
        shutil.copy2(tmp, alt)
        return alt

def row_height_for_text(text, min_height=30, px_per_line=14):
    """텍스트 줄 수에 따른 행 높이 자동 계산."""
    if not text:
        return min_height
    lines = text.count('\n') + 1
    return max(lines * px_per_line, min_height)
